__author__ = 'Shayhan'

import numpy as np
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import StandardScaler
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import os
from joblib import Parallel, delayed
import shutil
import math
import pickle
from constants import *
from datetime import timedelta

def evaluate_insulin_predictions(patient_info_list):
    df = pd.DataFrame(patient_info_list)
    # Add cutoff_time column
    df['cutoff_time'] = pd.to_datetime(df['pt_feeding_start_time']) - timedelta(days=1) + pd.to_timedelta(
        df['initial_bst_idx'], unit='m')

    df['BGL_post'] = df['target_bst']
    # df['BGL_post'] = df['initial_bst']

    # Initialize new columns
    df['group'] = ""
    df['correct'] = False

    # Assign groups
    df.loc[df['BGL_post'] <= 79, 'group'] = "Low"
    df.loc[(df['BGL_post'] >= 80) & (df['BGL_post'] <= 180), 'group'] = "Normal"
    df.loc[df['BGL_post'] >= 181, 'group'] = "High"

    groups = {
        "Low": df[df['BGL_post'] <= 79],
        "Normal": df[(df['BGL_post'] >= 80) & (df['BGL_post'] <= 180)],
        "High": df[df['BGL_post'] >= 181]
    }

    results = {}
    total_correct = 0
    total_count = 0

    for name, gdf in groups.items():
        if gdf.empty:
            results[name] = {"correct": 0, "total": 0, "accuracy": 0.0}
            continue

        IN_upper = gdf['IN_upper_hr']
        IN_lower = gdf['IN_lower_hr']
        IN_delta = gdf['IN_delta_hr']

        pred = gdf['predicted_insulin']
        actual = gdf['actual_insulin']

        if name == "Low":
            condition = (IN_lower <= pred) & (pred <= actual)
        elif name == "Normal":
            condition = (actual - IN_delta <= pred) & (pred <= actual + IN_delta)
        elif name == "High":
            condition = (actual <= pred) & (pred <= IN_upper)

        # Update 'correct' column in original df
        df.loc[gdf.index, 'correct'] = condition

        correct = condition.sum()

        # if name == "Low":
        #     correct = ((IN_lower <= pred) & (pred <= actual)).sum()
        # elif name == "Normal":
        #     correct = ((actual - IN_delta <= pred) & (pred <= actual + IN_delta)).sum()
        # elif name == "High":
        #     correct = ((actual <= pred) & (pred <= IN_upper)).sum()
        # else:
        #     correct = 0

        total = len(gdf)
        acc = (correct / total * 100) if total > 0 else 0.0
        results[name] = {"correct": correct, "total": total, "accuracy": acc}
        total_correct += correct
        total_count += total

    results["overall_accuracy"] = (total_correct / total_count * 100) if total_count > 0 else 0.0

    # --- Accuracy by hospital ---
    # Map 'KHNMC_V2' to 'KHNMC'
    df['hospital_grouped'] = df['hospital_name'].replace({'KHNMC_V2': 'KHNMC'})
    hospital_results = {}
    hospital_groups = df.groupby('hospital_grouped')

    for hospital, hdf in hospital_groups:
        total = len(hdf)
        correct = hdf['correct'].sum()
        acc = (correct / total * 100) if total > 0 else 0.0
        hospital_results[hospital] = {
            "correct": int(correct),
            "total": total,
            "accuracy": acc
        }

    results["by_hospital"] = hospital_results

    # Save with 'group' and 'correct' columns
    output_path = "results/main graph/current 898/expert.xlsx"
    df.to_excel(output_path, index=False)

    return results, df

def balace_train_data(X_train, y_train, bst_low, bst_high):
    indices = np.where((y_train >= bst_low) & (y_train <= bst_high))[0]
    X_balanced = np.concatenate((X_train, X_train[indices]), axis=0)
    y_balanced = np.concatenate((y_train, y_train[indices]), axis=0)
    return X_balanced, y_balanced



def generate_bell_curve(n, mean_index, std_dev):
    x = np.arange(n)
    return np.exp(-((x - mean_index) ** 2) / (2 * std_dev ** 2))

def bst_slope_calculation(cur_bst_idx, cur_bst, pre_bst_idx, pre_bst, future_bst_idx=0):
    # Perform linear interpolation to estimate the bst at end_portion_idx
    bst_slope = (cur_bst - pre_bst) / (
                cur_bst_idx - pre_bst_idx) if cur_bst_idx != pre_bst_idx else 0  # Avoid division by zero
    future_bst = pre_bst + bst_slope * (future_bst_idx - pre_bst_idx)
    if future_bst_idx == 0:
        return bst_slope
    else:
        return bst_slope, future_bst

def optimal_shift_window(final_bst_idx, final_bst, previous_bst_idx, previous_bst, calorie_seq, insulin_seq, duration):
    shif_interval = [0, 15, 30, 45, 60, 75, 90, 105, 120]
    min_rmse = float('inf')
    best_shift = 0
    best_predicted_bst = None
    for shift in shif_interval:
        shifted_previous_bst_idx = previous_bst_idx - shift
        shifted_final_bst_idx = final_bst_idx - shift
        calories = calorie_seq[shifted_previous_bst_idx:shifted_final_bst_idx].sum()
        insulin = insulin_seq[shifted_previous_bst_idx:shifted_final_bst_idx].sum()
        # igr = insulin / (calories + 0.01)
        predicted_bst = (previous_bst - 20) * math.exp(-0.01 * insulin) + (calories / 15) * math.exp(-0.005 * duration)
        # predicted_bst = gemini_calculate_final_bst(previous_bst, insulin, calories, duration)
        # predicted_bst = chat_gpt_calculate_final_bst(previous_bst, insulin, calories, duration, ICR=13, K=0.75)
        # predicted_bst = copilote_calculate_final_bst(previous_bst, insulin, calories, isf=50)
        rmse = math.sqrt((final_bst - predicted_bst) ** 2)
        if rmse < min_rmse:
            min_rmse = rmse
            best_shift = shift
            best_predicted_bst = predicted_bst
    return best_shift


def aggregate_minutes(data, period=10, method='max'):
    def non_zero_mean(arr):
        # Compute the mean of non-zero values
        non_zero_elements = arr[arr != 0]
        if non_zero_elements.size == 0:
            return 0  # Return 0 if there are no non-zero elements
        return np.mean(non_zero_elements)

    methods = {
        'sum': np.sum,
        'mean': np.mean,
        'max': np.max,
        'min': np.min,
        'std': np.std,
        'non_zero_mean': non_zero_mean  # Add the custom non-zero mean method
    }

    if method not in methods:
        raise ValueError(f"Invalid method '{method}'. Valid methods are {list(methods.keys())}.")

    if len(data.shape) == 3:
        # Reshape the second dimension to group every 'period' values together
        data = data.reshape(data.shape[0], data.shape[1] // period, period, data.shape[2])
    elif len(data.shape) == 2:
        data = data.reshape(data.shape[0], data.shape[1] // period, period)
    else:
        raise ValueError("Input data must be a 2D or 3D array.")

    # Compute the aggregation along the third dimension
    return np.apply_along_axis(methods[method], 2, data)


def reset_erros_folder(base_path):
    """
    Delete the 'erros' folder if it exists and make a copy of 'errors_empty' and rename it to 'erros'.

    Parameters:
    base_path (str): The base path where the folders are located.
    """
    erros_folder_path = os.path.join(base_path, "erros")
    errors_empty_folder_path = os.path.join(base_path, "errors_empty")
    new_erros_folder_path = os.path.join(base_path, "erros")

    # Delete the erros folder if it exists
    print(erros_folder_path)
    if os.path.exists(erros_folder_path):
        shutil.rmtree(erros_folder_path)
        print(f"Deleted the folder: {erros_folder_path}")

    # Make a copy of the errors_empty folder and rename it to erros
    shutil.copytree(errors_empty_folder_path, new_erros_folder_path)
    print(f"Copied and renamed the folder to: {new_erros_folder_path}")


def save_figure(fig, file_path):
    fig.savefig(file_path)
    plt.close(fig)




#     for i, pt in enumerate(patients_list):
#         plot_single_patient_data(folder_path, pt, patients_data[pt], error=errors[i] if len(errors) > 0 else None,
#         prediction_info=predictions_info[i] if len(predictions_info) > 0 else None, feeding_num=feeding_num, D_1_TIME=D_1_TIME, D_0_TIME=D_0_TIME, PATIENT_TYPE=PATIENT_TYPE)

    # Plot histogram of errors
    # plt.figure()
    # errors = np.array(errors)
    # n, bins, patches = plt.hist(errors, bins=20, color='skyblue', edgecolor='black')
    # plt.xlabel('Error Value (RMSE)')
    # plt.ylabel('Frequency')
    # plt.title('Histogram of Errors (Test Dataset)')
    # plt.xticks(np.arange(0, errors.max(), step=int(errors.max() / 20.)), rotation=45, ha='right',
    #            rotation_mode='anchor')
    #
    # # Adding the count on top of each bin
    # for count, patch in zip(n, patches):
    #     plt.text(patch.get_x() + patch.get_width() / 2, count, int(count), ha='center', va='bottom')
    # plt.tight_layout()
    #
    # # Save histogram
    # hist_path = os.path.join(folder_path, f"errors/fold_errors/{PATIENT_TYPE}/", f"fold_{fold}_RMSE_Hist.png")
    # plt.savefig(hist_path)
    # plt.close()


# def extract_features(crf_df):
#     bst_seq = np.array(crf_df['bst_sequence'].tolist())
#     insulin_samples = np.array(crf_df['insulin'].tolist())
#     calorie_samples = np.array(crf_df['calorie'].tolist())
#     ens_samples = np.array(crf_df['ens_distributed'].tolist())
#     insulin_i_samples = np.array(crf_df['insulin_i'].tolist())
#     insulin_h_samples = np.array(crf_df['insulin_h'].tolist())
#     insulin_i_iv_samples = np.array(crf_df['insulin_i_iv'].tolist())
#     tf_samples = calorie_samples - ens_samples
#     steroid_samples = np.array(crf_df['steroid'].tolist())
#     medication_samples = np.array(crf_df['medication'].tolist())
#     mask = medication_samples > 0
#     medication_samples[mask] = 1
#     all_features = {
#         'Insulin': insulin_samples,
#         'Steroid': steroid_samples,
#         'Total Calories': calorie_samples,
#         'TPN / Fluid': tf_samples,
#         # 'ENS': ens_samples,
#         'Medication': medication_samples,
#         'Insulin-I': insulin_i_samples,
#         'Insulin-H': insulin_h_samples,
#         'Insulin-IV': insulin_i_iv_samples
#     }
#     return bst_seq, all_features

def filter_patients(feeding_data, features, segments, condition_func):
    """
    Filters patients based on the given condition function and extracts features.

    Parameters:
    feeding_data (list): List of feeding data dictionaries.
    features (list): List of feature keys to extract.
    segments (list): Not used in the function (you might want to remove it if unnecessary).
    condition_func (function): A function that takes a feeding dictionary and returns a boolean.

    Returns:
    tuple: Arrays of X, y, pt_list, flt_feed.
    """
    X, y, action, pt_list, flt_feed, pt_cluster = [], [], [], [], [], []

    # for feed in feeding_data:
    for idx, feed in feeding_data.iterrows():
        if condition_func(feed):
            f = []
            for key in features:
                value = feed.get(key)
                f.append(float(value))

                # Handle missing features
                # if value is None:
                #     # Decide how you want to handle missing values
                #     # Here, we'll append NaN or a default value
                #     f.append(np.nan)
                # elif isinstance(value, list):
                #     # Decide on a fixed number of elements to extract
                #     # For example, extract the first N elements
                #     N = SEGMENT  # or any number appropriate for your data
                #     # Pad the list if it's shorter than N
                #     padded_value = (value + [np.nan]*N)[:N]
                #     f.extend(padded_value)
                # else:
                #     f.append(float(value))
            X.append(f)
            y.append(feed['target_bst'])
            action.append((feed['complete_insulin'] * 60.) / feed['duration'])
            # action.append((feed['complete_insulin_a'] * 60.) / feed['duration'])
            # action.append((feed['complete_insulin_h'] * 60.) / feed['duration'])
            # action.append(((feed['complete_insulin'] - feed['complete_insulin_i']) * 60.) / feed['duration'])
            # action.append((feed['s_insulin_sum'] * 60.) / feed['duration'])
            # action.append(feed['complete_insulin'])
            # y.append(feed['bst_change'])
            pt_list.append(feed["pt_index"])
            flt_feed.append(feed)
            # pt_cluster.append(feed["cluster"])
            pt_cluster.append(feed["hospital_name"])

    # Convert to NumPy arrays
    X = np.array(X)
    y = np.array(y)
    action = np.array(action)

    # Handle any remaining inconsistencies
    # For example, remove samples with NaN values if necessary
    valid_indices = ~np.isnan(X).any(axis=1)
    # print(valid_indices.shape)
    X = X[valid_indices]
    y = y[valid_indices]
    action = action[valid_indices]
    pt_list = np.array(pt_list)[valid_indices]
    flt_feed = [flt_feed[i] for i in range(len(flt_feed)) if valid_indices[i]]
    pt_cluster = np.array(pt_cluster)[valid_indices]

    return X, y, action, pt_list, flt_feed, pt_cluster


# def filter_patients(feeding_data, features, segments, condition_func):
#     """
#     Filters patients based on the given condition function and extracts features.
#
#     Parameters:
#     feeding_data (list): List of feeding data dictionaries.
#     features (list): List of feature keys to extract.
#     segments (list): Not used in the function (you might want to remove it if unnecessary).
#     condition_func (function): A function that takes a feeding dictionary and returns a boolean.
#
#     Returns:
#     tuple: Arrays of X, y, pt_list, flt_feed.
#     """
#     X, y, pt_list, flt_feed = [], [], [], []
#
#     for feed in feeding_data:
#         if condition_func(feed):
#             f = []
#             for key in features:
#                 value = feed.get(key)
#
#                 # Handle missing features
#                 if value is None:
#                     # Decide how you want to handle missing values
#                     # Here, we'll append NaN or a default value
#                     f.append(np.nan)
#                 elif isinstance(value, list):
#                     # Decide on a fixed number of elements to extract
#                     # For example, extract the first N elements
#                     N = SEGMENT  # or any number appropriate for your data
#                     # Pad the list if it's shorter than N
#                     padded_value = (value + [np.nan]*N)[:N]
#                     f.extend(padded_value)
#                 else:
#                     f.append(float(value))
#             X.append(f)
#             y.append(feed['target_bst'])
#             # y.append(feed['bst_change'])
#             pt_list.append(feed["pt_index"])
#             flt_feed.append(feed)
#
#     # Convert to NumPy arrays
#     X = np.array(X)
#     y = np.array(y)
#
#     # Handle any remaining inconsistencies
#     # For example, remove samples with NaN values if necessary
#     valid_indices = ~np.isnan(X).any(axis=1)
#     X = X[valid_indices]
#     y = y[valid_indices]
#     pt_list = np.array(pt_list)[valid_indices]
#     flt_feed = [flt_feed[i] for i in range(len(flt_feed)) if valid_indices[i]]
#
#     return X, y, pt_list, flt_feed



def cluster_patients_crf(crf_df, n_clusters=3, sequence_size=1400, seed_value=99):
    new_columns = pd.DataFrame({
        'bst_non_zero_mean': crf_df['bst_sequence'].apply(
            lambda x: np.mean(x[:sequence_size][x[:sequence_size] != 0]) if len(
                x[:sequence_size][x[:sequence_size] != 0]) > 0 else 0),
        'insulin_non_zero_mean': crf_df['insulin'].apply(
            lambda x: np.mean(x[:sequence_size][x[:sequence_size] != 0]) if len(
                x[:sequence_size][x[:sequence_size] != 0]) > 0 else 0),
        'calorie_non_zero_mean': crf_df['calorie'].apply(
            lambda x: np.mean(x[:sequence_size][x[:sequence_size] != 0]) if len(
                x[:sequence_size][x[:sequence_size] != 0]) > 0 else 0)
    })

    crf_df = pd.concat([crf_df, new_columns], axis=1)
    crf_df['bmi'] = crf_df['Bwt']/crf_df['Ht']
    crf_df['Gender'] = crf_df['Gender'].apply(lambda x: 1 if x == 'M' else 0)
    crf_df['igr'] = crf_df['insulin_non_zero_mean'] / (crf_df['calorie_non_zero_mean'] + 0.01)
    # clustering_columns = ['bst_non_zero_mean', 'insulin_non_zero_mean', 'calorie_non_zero_mean']
    clustering_columns = ['bst_non_zero_mean', 'Age', 'Gender', 'bmi', 'igr']

    # "age": pt_data['age'],
    # "gender": pt_data['gender'],
    # "hgt": pt_data['hgt'],
    # "bwt": pt_data['bwt'],

    # Fill missing values with the mean of each column
    crf_df[clustering_columns] = crf_df[clustering_columns].apply(lambda x: x.fillna(x.mean()), axis=0)

    # Compute IGR (Insulin to Glucose Ratio)


    # Select data for clustering
    # clustering_data = crf_df[['bst_non_zero_mean', 'igr']]
    clustering_data = crf_df[clustering_columns]

    # Standardize the data
    scaler = StandardScaler()
    clustering_data = scaler.fit_transform(clustering_data)

    # Perform K-Means clustering
    kmeans = KMeans(n_clusters=n_clusters, random_state=seed_value)
    crf_df['cluster'] = kmeans.fit_predict(clustering_data)

    # # Visualization
    plt.scatter(clustering_data[:, 0], clustering_data[:, -1], c=crf_df['cluster'], cmap='viridis')
    plt.xlabel('AVG BGL')
    plt.ylabel('Insulin to Glucose Ratio (IGR)')
    plt.title('Clustering Results')
    plt.colorbar(label='Cluster')
    plt.show()

    # 3D Scatter plot
    # fig = plt.figure(figsize=(10, 8))
    # ax = fig.add_subplot(111, projection='3d')
    # # Scatter points
    # scatter = ax.scatter(crf_df['bst_non_zero_mean'], crf_df['Age'], crf_df['bmi'], c=crf_df['cluster'], cmap='viridis', s=50)
    # # Add axis labels
    # ax.set_xlabel('Avg. BGL')
    # ax.set_ylabel('Age')
    # ax.set_zlabel('BMI')
    # ax.set_title('3D Cluster Visualization')
    #
    # # Add color bar
    # plt.colorbar(scatter, ax=ax, label='Cluster')
    # plt.show()

    # Count the number of data points in each cluster and Display the counts
    # cluster_counts = crf_df['cluster'].value_counts()
    # print(cluster_counts)

    return crf_df

# def cluster_patients(crf_df, n_clusters=3, sequence_size=1400, seed_value=99):
#     new_columns = pd.DataFrame({
#         'bst_non_zero_mean': crf_df['bst_sequence'].apply(
#             lambda x: np.mean(x[:sequence_size][x[:sequence_size] != 0]) if len(
#                 x[:sequence_size][x[:sequence_size] != 0]) > 0 else 0),
#         'insulin_non_zero_mean': crf_df['insulin'].apply(
#             lambda x: np.mean(x[:sequence_size][x[:sequence_size] != 0]) if len(
#                 x[:sequence_size][x[:sequence_size] != 0]) > 0 else 0),
#         'calorie_non_zero_mean': crf_df['calorie'].apply(
#             lambda x: np.mean(x[:sequence_size][x[:sequence_size] != 0]) if len(
#                 x[:sequence_size][x[:sequence_size] != 0]) > 0 else 0)
#     })
#
#     crf_df = pd.concat([crf_df, new_columns], axis=1)
#     clustering_columns = ['bst_non_zero_mean', 'insulin_non_zero_mean', 'calorie_non_zero_mean']
#
#     # Fill missing values with the mean of each column
#     crf_df[clustering_columns] = crf_df[clustering_columns].apply(lambda x: x.fillna(x.mean()), axis=0)
#
#     # Compute IGR (Insulin to Glucose Ratio)
#     crf_df['igr'] = crf_df['insulin_non_zero_mean'] / (crf_df['calorie_non_zero_mean'] + 0.01)
#
#     # Select data for clustering
#     clustering_data = crf_df[['bst_non_zero_mean', 'igr']]
#
#     # Standardize the data
#     scaler = StandardScaler()
#     clustering_data = scaler.fit_transform(clustering_data)
#
#     # Perform K-Means clustering
#     kmeans = KMeans(n_clusters=n_clusters, random_state=seed_value)
#     crf_df['cluster'] = kmeans.fit_predict(clustering_data)
#
#     # # Visualization
#     # plt.scatter(clustering_data[:, 0], clustering_data[:, 1], c=crf_df['cluster'], cmap='viridis')
#     # plt.xlabel('BST Non-Zero Mean')
#     # plt.ylabel('Insulin to Glucose Ratio (IGR)')
#     # plt.title('Clustering Results')
#     # plt.colorbar(label='Cluster')
#     # plt.show()
#
#     return crf_df

def display_evaluation_summary(rmses, mses, maes, X_train, X_val, X_test, feeding_num, patient_type, bst_range, target_min, target_max,
                             best_parameters=None, SHOW_BEST_PARAM=False, SHORT_MESSAGE=True):

    # Calculate averages
    average_rmse, average_mse, average_mae = map(np.mean, [rmses, mses, maes])

    # Print relevant information
    if SHORT_MESSAGE:
        print(feeding_num)
        print(patient_type)
        print(bst_range)
        print(f"Target BST: [{target_min} - {target_max}]")
        print(f'Total: {X_train.shape[0] + X_val.shape[0] + X_test.shape[0]}')
        print(f'Train: {X_train.shape[0]}')
        print(f'Val: {X_val.shape[0]}')
        print(f'Test: {X_test.shape[0]}')

    else:
        print(f'Feeding number: {feeding_num}')
        print(f'Patient type: {patient_type}')
        print(f'BST range: {bst_range}')
        print(f'Train set size: {X_train.shape[0]}')
        print(f'Validation set size: {X_val.shape[0]}')
        print(f'Test set size: {X_test.shape[0]}')

    # Print results for each fold
    for i, rmse in enumerate(rmses):
        if SHOW_BEST_PARAM and best_parameters is not None:
            print(f'Fold {i + 1}: RMSE: {rmse:.2f} || Best Parameters: {best_parameters[i]}')
            # print(f'Fold {i + 1}: RMSE: {rmse:.2f} || Best Parameters: {best_parameters}')
        else:
            print(f'Fold {i + 1}: RMSE: {rmse:.2f}')

    # Print average metrics
    if SHORT_MESSAGE:
        print(f'Avg. RMSE: {average_rmse:.2f}')
    else:
        print(f'Average RMSE: {average_rmse:.2f}')
    # print(f'Average MAE: {average_mae:.2f}')
    # print(f'Average MSE: {average_mse:.2f}')

def initialize_ex_results():
    ex_results = {}
    feeding_numbers = [1, 2, 3]
    patient_types = ["without", "with"]
    bst_ranges = [[0, 70], [71, 200], [201, 300], [301, 400], [401, 701]]
    for feeding_num in feeding_numbers:
        for patient_type in patient_types:
            for bst_range in bst_ranges:
                for fold in range(1, 6):
                    r = f"{patient_type}_Steroid-Feeding_{feeding_num}-BST_RANGE_{bst_range}-FOLD_{fold}"
                    ex_results[r] = ""
    return ex_results


import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def store_experiment_results(
    filename,
    model,
    dataset,
    hospital,
    patient_type,
    bst_range,
    feeding_no,
    wrmse,
    features,
    fold_rmses
):
    """
    Appends a new experiment result to the specified Excel (.xlsx) file.

    Parameters:
    - filename (str): The name of the Excel file.
    - model (str): The model used in the experiment.
    - dataset (str): The dataset used.
    - hospital (str): The hospital data source.
    - patient_type (str): The type of patients.
    - bst_range (str): The BST range.
    - feeding_no (int): The feeding number.
    - wrmse (float): The weighted RMSE of the experiment.
    - fold_rmses (dict): A dictionary containing RMSEs for each fold and feature set.
      Example:
      {
          'without_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_1': value,
          'w/o_F1_FOLD_2_RMSE': value,
          ...
          'w/o_F2_FOLD_5_RMSE': value
      }
    """
    headers = [
        'Date', 'Experiment No', 'Model', 'Dataset', 'Hospital', 'Patient Type',
        'BST Range', 'Feeding No.', 'wRMSE', 'Features',
        "without_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_1",
        "without_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_2",
        "without_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_3",
        "without_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_4",
        "without_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_5",
        "without_Steroid-Feeding_1-BST_RANGE_[0, 70]-AVERAGE",
        "without_Steroid-Feeding_1-BST_RANGE_[71, 200]-FOLD_1",
        "without_Steroid-Feeding_1-BST_RANGE_[71, 200]-FOLD_2",
        "without_Steroid-Feeding_1-BST_RANGE_[71, 200]-FOLD_3",
        "without_Steroid-Feeding_1-BST_RANGE_[71, 200]-FOLD_4",
        "without_Steroid-Feeding_1-BST_RANGE_[71, 200]-FOLD_5",
        "without_Steroid-Feeding_1-BST_RANGE_[71, 200]-AVERAGE",
        "without_Steroid-Feeding_1-BST_RANGE_[201, 300]-FOLD_1",
        "without_Steroid-Feeding_1-BST_RANGE_[201, 300]-FOLD_2",
        "without_Steroid-Feeding_1-BST_RANGE_[201, 300]-FOLD_3",
        "without_Steroid-Feeding_1-BST_RANGE_[201, 300]-FOLD_4",
        "without_Steroid-Feeding_1-BST_RANGE_[201, 300]-FOLD_5",
        "without_Steroid-Feeding_1-BST_RANGE_[201, 300]-AVERAGE",
        "without_Steroid-Feeding_1-BST_RANGE_[301, 400]-FOLD_1",
        "without_Steroid-Feeding_1-BST_RANGE_[301, 400]-FOLD_2",
        "without_Steroid-Feeding_1-BST_RANGE_[301, 400]-FOLD_3",
        "without_Steroid-Feeding_1-BST_RANGE_[301, 400]-FOLD_4",
        "without_Steroid-Feeding_1-BST_RANGE_[301, 400]-FOLD_5",
        "without_Steroid-Feeding_1-BST_RANGE_[301, 400]-AVERAGE",
        "without_Steroid-Feeding_1-BST_RANGE_[401, 701]-FOLD_1",
        "without_Steroid-Feeding_1-BST_RANGE_[401, 701]-FOLD_2",
        "without_Steroid-Feeding_1-BST_RANGE_[401, 701]-FOLD_3",
        "without_Steroid-Feeding_1-BST_RANGE_[401, 701]-FOLD_4",
        "without_Steroid-Feeding_1-BST_RANGE_[401, 701]-FOLD_5",
        "without_Steroid-Feeding_1-BST_RANGE_[401, 701]-AVERAGE",
        "with_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_1",
        "with_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_2",
        "with_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_3",
        "with_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_4",
        "with_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_5",
        "with_Steroid-Feeding_1-BST_RANGE_[0, 70]-AVERAGE",
        "with_Steroid-Feeding_1-BST_RANGE_[71, 200]-FOLD_1",
        "with_Steroid-Feeding_1-BST_RANGE_[71, 200]-FOLD_2",
        "with_Steroid-Feeding_1-BST_RANGE_[71, 200]-FOLD_3",
        "with_Steroid-Feeding_1-BST_RANGE_[71, 200]-FOLD_4",
        "with_Steroid-Feeding_1-BST_RANGE_[71, 200]-FOLD_5",
        "with_Steroid-Feeding_1-BST_RANGE_[71, 200]-AVERAGE",
        "with_Steroid-Feeding_1-BST_RANGE_[201, 300]-FOLD_1",
        "with_Steroid-Feeding_1-BST_RANGE_[201, 300]-FOLD_2",
        "with_Steroid-Feeding_1-BST_RANGE_[201, 300]-FOLD_3",
        "with_Steroid-Feeding_1-BST_RANGE_[201, 300]-FOLD_4",
        "with_Steroid-Feeding_1-BST_RANGE_[201, 300]-FOLD_5",
        "with_Steroid-Feeding_1-BST_RANGE_[201, 300]-AVERAGE",
        "with_Steroid-Feeding_1-BST_RANGE_[301, 400]-FOLD_1",
        "with_Steroid-Feeding_1-BST_RANGE_[301, 400]-FOLD_2",
        "with_Steroid-Feeding_1-BST_RANGE_[301, 400]-FOLD_3",
        "with_Steroid-Feeding_1-BST_RANGE_[301, 400]-FOLD_4",
        "with_Steroid-Feeding_1-BST_RANGE_[301, 400]-FOLD_5",
        "with_Steroid-Feeding_1-BST_RANGE_[301, 400]-AVERAGE",
        "with_Steroid-Feeding_1-BST_RANGE_[401, 701]-FOLD_1",
        "with_Steroid-Feeding_1-BST_RANGE_[401, 701]-FOLD_2",
        "with_Steroid-Feeding_1-BST_RANGE_[401, 701]-FOLD_3",
        "with_Steroid-Feeding_1-BST_RANGE_[401, 701]-FOLD_4",
        "with_Steroid-Feeding_1-BST_RANGE_[401, 701]-FOLD_5",
        "with_Steroid-Feeding_1-BST_RANGE_[401, 701]-AVERAGE",
        "without_Steroid-Feeding_2-BST_RANGE_[0, 70]-FOLD_1",
        "without_Steroid-Feeding_2-BST_RANGE_[0, 70]-FOLD_2",
        "without_Steroid-Feeding_2-BST_RANGE_[0, 70]-FOLD_3",
        "without_Steroid-Feeding_2-BST_RANGE_[0, 70]-FOLD_4",
        "without_Steroid-Feeding_2-BST_RANGE_[0, 70]-FOLD_5",
        "without_Steroid-Feeding_2-BST_RANGE_[0, 70]-AVERAGE",
        "without_Steroid-Feeding_2-BST_RANGE_[71, 200]-FOLD_1",
        "without_Steroid-Feeding_2-BST_RANGE_[71, 200]-FOLD_2",
        "without_Steroid-Feeding_2-BST_RANGE_[71, 200]-FOLD_3",
        "without_Steroid-Feeding_2-BST_RANGE_[71, 200]-FOLD_4",
        "without_Steroid-Feeding_2-BST_RANGE_[71, 200]-FOLD_5",
        "without_Steroid-Feeding_2-BST_RANGE_[71, 200]-AVERAGE",
        "without_Steroid-Feeding_2-BST_RANGE_[201, 300]-FOLD_1",
        "without_Steroid-Feeding_2-BST_RANGE_[201, 300]-FOLD_2",
        "without_Steroid-Feeding_2-BST_RANGE_[201, 300]-FOLD_3",
        "without_Steroid-Feeding_2-BST_RANGE_[201, 300]-FOLD_4",
        "without_Steroid-Feeding_2-BST_RANGE_[201, 300]-FOLD_5",
        "without_Steroid-Feeding_2-BST_RANGE_[201, 300]-AVERAGE",
        "without_Steroid-Feeding_2-BST_RANGE_[301, 400]-FOLD_1",
        "without_Steroid-Feeding_2-BST_RANGE_[301, 400]-FOLD_2",
        "without_Steroid-Feeding_2-BST_RANGE_[301, 400]-FOLD_3",
        "without_Steroid-Feeding_2-BST_RANGE_[301, 400]-FOLD_4",
        "without_Steroid-Feeding_2-BST_RANGE_[301, 400]-FOLD_5",
        "without_Steroid-Feeding_2-BST_RANGE_[301, 400]-AVERAGE",
        "without_Steroid-Feeding_2-BST_RANGE_[401, 701]-FOLD_1",
        "without_Steroid-Feeding_2-BST_RANGE_[401, 701]-FOLD_2",
        "without_Steroid-Feeding_2-BST_RANGE_[401, 701]-FOLD_3",
        "without_Steroid-Feeding_2-BST_RANGE_[401, 701]-FOLD_4",
        "without_Steroid-Feeding_2-BST_RANGE_[401, 701]-FOLD_5",
        "without_Steroid-Feeding_2-BST_RANGE_[401, 701]-AVERAGE",
        "with_Steroid-Feeding_2-BST_RANGE_[0, 70]-FOLD_1",
        "with_Steroid-Feeding_2-BST_RANGE_[0, 70]-FOLD_2",
        "with_Steroid-Feeding_2-BST_RANGE_[0, 70]-FOLD_3",
        "with_Steroid-Feeding_2-BST_RANGE_[0, 70]-FOLD_4",
        "with_Steroid-Feeding_2-BST_RANGE_[0, 70]-FOLD_5",
        "with_Steroid-Feeding_2-BST_RANGE_[0, 70]-AVERAGE",
        "with_Steroid-Feeding_2-BST_RANGE_[71, 200]-FOLD_1",
        "with_Steroid-Feeding_2-BST_RANGE_[71, 200]-FOLD_2",
        "with_Steroid-Feeding_2-BST_RANGE_[71, 200]-FOLD_3",
        "with_Steroid-Feeding_2-BST_RANGE_[71, 200]-FOLD_4",
        "with_Steroid-Feeding_2-BST_RANGE_[71, 200]-FOLD_5",
        "with_Steroid-Feeding_2-BST_RANGE_[71, 200]-AVERAGE",
        "with_Steroid-Feeding_2-BST_RANGE_[201, 300]-FOLD_1",
        "with_Steroid-Feeding_2-BST_RANGE_[201, 300]-FOLD_2",
        "with_Steroid-Feeding_2-BST_RANGE_[201, 300]-FOLD_3",
        "with_Steroid-Feeding_2-BST_RANGE_[201, 300]-FOLD_4",
        "with_Steroid-Feeding_2-BST_RANGE_[201, 300]-FOLD_5",
        "with_Steroid-Feeding_2-BST_RANGE_[201, 300]-AVERAGE",
        "with_Steroid-Feeding_2-BST_RANGE_[301, 400]-FOLD_1",
        "with_Steroid-Feeding_2-BST_RANGE_[301, 400]-FOLD_2",
        "with_Steroid-Feeding_2-BST_RANGE_[301, 400]-FOLD_3",
        "with_Steroid-Feeding_2-BST_RANGE_[301, 400]-FOLD_4",
        "with_Steroid-Feeding_2-BST_RANGE_[301, 400]-FOLD_5",
        "with_Steroid-Feeding_2-BST_RANGE_[301, 400]-AVERAGE",
        "with_Steroid-Feeding_2-BST_RANGE_[401, 701]-FOLD_1",
        "with_Steroid-Feeding_2-BST_RANGE_[401, 701]-FOLD_2",
        "with_Steroid-Feeding_2-BST_RANGE_[401, 701]-FOLD_3",
        "with_Steroid-Feeding_2-BST_RANGE_[401, 701]-FOLD_4",
        "with_Steroid-Feeding_2-BST_RANGE_[401, 701]-FOLD_5",
        "with_Steroid-Feeding_2-BST_RANGE_[401, 701]-AVERAGE",
        "without_Steroid-Feeding_3-BST_RANGE_[0, 70]-FOLD_1",
        "without_Steroid-Feeding_3-BST_RANGE_[0, 70]-FOLD_2",
        "without_Steroid-Feeding_3-BST_RANGE_[0, 70]-FOLD_3",
        "without_Steroid-Feeding_3-BST_RANGE_[0, 70]-FOLD_4",
        "without_Steroid-Feeding_3-BST_RANGE_[0, 70]-FOLD_5",
        "without_Steroid-Feeding_3-BST_RANGE_[0, 70]-AVERAGE",
        "without_Steroid-Feeding_3-BST_RANGE_[71, 200]-FOLD_1",
        "without_Steroid-Feeding_3-BST_RANGE_[71, 200]-FOLD_2",
        "without_Steroid-Feeding_3-BST_RANGE_[71, 200]-FOLD_3",
        "without_Steroid-Feeding_3-BST_RANGE_[71, 200]-FOLD_4",
        "without_Steroid-Feeding_3-BST_RANGE_[71, 200]-FOLD_5",
        "without_Steroid-Feeding_3-BST_RANGE_[71, 200]-AVERAGE",
        "without_Steroid-Feeding_3-BST_RANGE_[201, 300]-FOLD_1",
        "without_Steroid-Feeding_3-BST_RANGE_[201, 300]-FOLD_2",
        "without_Steroid-Feeding_3-BST_RANGE_[201, 300]-FOLD_3",
        "without_Steroid-Feeding_3-BST_RANGE_[201, 300]-FOLD_4",
        "without_Steroid-Feeding_3-BST_RANGE_[201, 300]-FOLD_5",
        "without_Steroid-Feeding_3-BST_RANGE_[201, 300]-AVERAGE",
        "without_Steroid-Feeding_3-BST_RANGE_[301, 400]-FOLD_1",
        "without_Steroid-Feeding_3-BST_RANGE_[301, 400]-FOLD_2",
        "without_Steroid-Feeding_3-BST_RANGE_[301, 400]-FOLD_3",
        "without_Steroid-Feeding_3-BST_RANGE_[301, 400]-FOLD_4",
        "without_Steroid-Feeding_3-BST_RANGE_[301, 400]-FOLD_5",
        "without_Steroid-Feeding_3-BST_RANGE_[301, 400]-AVERAGE",
        "without_Steroid-Feeding_3-BST_RANGE_[401, 701]-FOLD_1",
        "without_Steroid-Feeding_3-BST_RANGE_[401, 701]-FOLD_2",
        "without_Steroid-Feeding_3-BST_RANGE_[401, 701]-FOLD_3",
        "without_Steroid-Feeding_3-BST_RANGE_[401, 701]-FOLD_4",
        "without_Steroid-Feeding_3-BST_RANGE_[401, 701]-FOLD_5",
        "without_Steroid-Feeding_3-BST_RANGE_[401, 701]-AVERAGE",
        "with_Steroid-Feeding_3-BST_RANGE_[0, 70]-FOLD_1",
        "with_Steroid-Feeding_3-BST_RANGE_[0, 70]-FOLD_2",
        "with_Steroid-Feeding_3-BST_RANGE_[0, 70]-FOLD_3",
        "with_Steroid-Feeding_3-BST_RANGE_[0, 70]-FOLD_4",
        "with_Steroid-Feeding_3-BST_RANGE_[0, 70]-FOLD_5",
        "with_Steroid-Feeding_3-BST_RANGE_[0, 70]-AVERAGE",
        "with_Steroid-Feeding_3-BST_RANGE_[71, 200]-FOLD_1",
        "with_Steroid-Feeding_3-BST_RANGE_[71, 200]-FOLD_2",
        "with_Steroid-Feeding_3-BST_RANGE_[71, 200]-FOLD_3",
        "with_Steroid-Feeding_3-BST_RANGE_[71, 200]-FOLD_4",
        "with_Steroid-Feeding_3-BST_RANGE_[71, 200]-FOLD_5",
        "with_Steroid-Feeding_3-BST_RANGE_[71, 200]-AVERAGE",
        "with_Steroid-Feeding_3-BST_RANGE_[201, 300]-FOLD_1",
        "with_Steroid-Feeding_3-BST_RANGE_[201, 300]-FOLD_2",
        "with_Steroid-Feeding_3-BST_RANGE_[201, 300]-FOLD_3",
        "with_Steroid-Feeding_3-BST_RANGE_[201, 300]-FOLD_4",
        "with_Steroid-Feeding_3-BST_RANGE_[201, 300]-FOLD_5",
        "with_Steroid-Feeding_3-BST_RANGE_[201, 300]-AVERAGE",
        "with_Steroid-Feeding_3-BST_RANGE_[301, 400]-FOLD_1",
        "with_Steroid-Feeding_3-BST_RANGE_[301, 400]-FOLD_2",
        "with_Steroid-Feeding_3-BST_RANGE_[301, 400]-FOLD_3",
        "with_Steroid-Feeding_3-BST_RANGE_[301, 400]-FOLD_4",
        "with_Steroid-Feeding_3-BST_RANGE_[301, 400]-FOLD_5",
        "with_Steroid-Feeding_3-BST_RANGE_[301, 400]-AVERAGE",
        "with_Steroid-Feeding_3-BST_RANGE_[401, 701]-FOLD_1",
        "with_Steroid-Feeding_3-BST_RANGE_[401, 701]-FOLD_2",
        "with_Steroid-Feeding_3-BST_RANGE_[401, 701]-FOLD_3",
        "with_Steroid-Feeding_3-BST_RANGE_[401, 701]-FOLD_4",
        "with_Steroid-Feeding_3-BST_RANGE_[401, 701]-FOLD_5",
        "with_Steroid-Feeding_3-BST_RANGE_[401, 701]-AVERAGE"
    ]

    # Check if the Excel file already exists
    if not os.path.isfile(filename):
        # If file doesn't exist, create a new workbook and add headers
        wb = Workbook()
        ws = wb.active
        ws.title = 'Experiment Results'
        ws.append(headers)  # Write the headers
        experiment_no = 0
    else:
        # If file exists, load the workbook and get the sheet
        wb = load_workbook(filename)
        ws = wb.active

        # Read the last experiment number from the existing data
        experiment_no = 0
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            experiment_no = max(experiment_no, row[1] or 0)  # Experiment No. is in the second column
        experiment_no += 1  # Increment the experiment number for the new entry

    # Prepare the data row
    data_row = [
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),  # Date
        experiment_no,                                  # Experiment No
        model,                                          # Model
        dataset,                                        # Dataset
        hospital,                                       # Hospital
        patient_type,                                   # Patient Type
        bst_range,                                      # BST Range
        feeding_no,                                     # Feeding No
        wrmse,                                          # wRMSE
        features                                        # Features
    ]

    # # Add fold RMSEs to the data row
    # for key in headers:
    #     if key.startswith('with'):
    #         data_row[key] = fold_rmses.get(key, '')
    # # Append the data row to the CSV file
    # with open(filename, 'a', newline='') as csvfile:
    #     writer = csv.DictWriter(csvfile, fieldnames=headers)
    #     writer.writerow(data_row)

    # Append fold RMSEs based on the headers
    for header in headers[10:]:  # Skip the first 10 columns (fixed info columns)
        data_row.append(fold_rmses.get(header, ''))  # Fill in the RMSEs or leave empty

    # Append the new row to the worksheet
    ws.append(data_row)

    # Save the workbook
    wb.save(filename)

    print(f"Experiment {experiment_no} results appended to '{filename}'.")


# Example usage
if __name__ == "__main__":
    # Define the fold RMSEs
    fold_rmses = {
        'without_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_1': 1.23,
        'without_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_2': 1.45,
        'without_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_3': 1.34,
        'without_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_4': 1.29,
        'without_Steroid-Feeding_1-BST_RANGE_[0, 70]-FOLD_5': 1.31,
        'without_Steroid-Feeding_1-BST_RANGE_[0, 70]-AVERAGE': 1.32,
    }

    # Call the function to store results in an Excel file
    store_experiment_results_xlsx(
        filename='experiment_results.xlsx',
        model='XGBoost',
        dataset='Dataset Name',
        hospital='General Hospital',
        patient_type='Type A',
        bst_range='70-150',
        feeding_no=3,
        wrmse=1.33,
        features='Feature Set X',
        fold_rmses=fold_rmses
    )

import numpy as np


def filter_nan_rows(X, flt_pts, min_rows=20):
    """
    This function filters out rows with NaN values from the input array X and
    prints the row indices and column indices of NaN values. It skips processing
    if there are fewer than min_rows rows after filtering.

    Parameters:
    - X: Input numpy array
    - flt_pts: List of additional information corresponding to rows in X
    - min_rows: Minimum number of rows required to proceed (default: 20)

    Returns:
    - X_filtered: The filtered array with rows containing NaNs removed
    """
    if X.shape[0] < min_rows:
        print("Skipping due to insufficient data")
        return None  # Return None if insufficient data initially

    rows_to_keep = []  # List to store rows without NaN

    for i in range(X.shape[0]):  # Loop over rows
        if np.isnan(X[i, :]).any():  # Check if the row contains any NaN values
            nan_columns = np.where(np.isnan(X[i, :]))[0]  # Find column indices of NaN
            print(f"Row {i} contains NaN in columns {nan_columns}")
            print(f'{flt_pts[i]}')
        else:
            rows_to_keep.append(i)  # Keep track of rows without NaN

    # Discard rows with NaN values by selecting only the rows to keep
    X_filtered = X[rows_to_keep, :]

    if X_filtered.shape[0] < min_rows:
        print("Skipping due to insufficient data after removing NaNs")
        return None  # Return None if insufficient data after filtering

    return X_filtered  # Return the filtered array if enough rows remain

# Example usage:
# X_filtered = filter_nan_rows(X, flt_pts)
# if X_filtered is not None:
#     # Proceed with further processing
